#!/usr/bin/env python3
"""
按运营主体拆表工具
将待拆表格按码表中指定对接人的运营主体拆分

使用方法：
  python split_by_zhuti.py           # 正常运行
  python split_by_zhuti.py --config  # 重新配置
"""

import os
import sys
import zipfile
from datetime import datetime
from pathlib import Path
from copy import copy

import pandas as pd
import yaml
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ========================================
# 路径配置（自动检测，无需修改）
# ========================================
SKILL_DIR = Path(__file__).parent.parent  # skill 目录
CONFIG_PATH = SKILL_DIR / "assets" / "config.yaml"
CONFIG_EXAMPLE_PATH = SKILL_DIR / "assets" / "config.yaml.example"


def interactive_setup():
    """交互式配置"""
    print("\n" + "=" * 60)
    print("按运营主体拆表工具 - 首次配置")
    print("=" * 60)

    # 码表说明
    print("\n【重要】使用本工具前请确保已准备好码表文件：")
    print("  码表是一份 Excel 文件，用于建立 品牌+城市 → 运营主体 的映射关系")
    print("")
    print("  码表必须包含以下列：")
    print("    - 品牌（或品牌名称）")
    print("    - 城市（或城市名称）")
    print("    - 运营主体（或新-运营主体）")
    print("    - 对接人")
    print("")
    print("  推荐放置位置：")
    print("    config/lx码表.xlsx")
    print("  或放在项目任意位置，配置时指定路径即可")
    print("-" * 60)

    print("\n检测到未配置，请提供以下信息：\n")

    # 1. 项目根目录
    current_dir = Path.cwd()
    print(f"1. 项目根目录")
    print(f"   当前目录: {current_dir}")
    user_input = input("   回车使用当前目录，或输入路径: ").strip()
    project_root = Path(user_input) if user_input else current_dir
    project_root = project_root.resolve()

    # 2. 码表路径
    print(f"\n2. 码表文件路径")
    print("   推荐: config/lx码表.xlsx")
    print("   或绝对路径: /Users/xxx/projects/my-project/config/lx码表.xlsx")
    user_input = input("   输入码表路径（相对或绝对路径）: ").strip()
    if not user_input:
        print("   ❌ 码表路径不能为空")
        return None
    mabiao_path = Path(user_input)
    if not mabiao_path.is_absolute():
        mabiao_path = project_root / mabiao_path

    # 3. 码表 Sheet
    print(f"\n3. 码表所在 Sheet")
    print("   如果码表在第一个 Sheet，直接回车即可")
    user_input = input("   输入 Sheet 名称或索引（默认: 第一个Sheet）: ").strip()
    mabiao_sheet = user_input if user_input else None

    # 4. 对接人
    print(f"\n4. 对接人")
    print("   需要拆出哪个对接人的表格？（多个用逗号分隔，回车拆出所有）")
    user_input = input("   对接人: ").strip()
    duijieren = user_input if user_input else "全部"

    # 5. 工作目录
    print(f"\n5. 工作目录")
    print("   目录下需包含：待拆表/、已拆表/、原表存档/ 三个子目录")
    user_input = input("   输入工作目录名称（默认: p-主体拆表）: ").strip()
    work_dir_name = user_input if user_input else "p-主体拆表"
    work_dir = project_root / work_dir_name

    # 确认配置
    print("\n" + "-" * 60)
    print("配置确认：")
    print(f"  项目根目录: {project_root}")
    print(f"  码表路径: {mabiao_path}")
    if mabiao_sheet:
        print(f"  码表Sheet: {mabiao_sheet}")
    print(f"  对接人: {duijieren}")
    print(f"  工作目录: {work_dir}")
    print("-" * 60)

    user_input = input("\n确认保存配置？(Y/n): ").strip().lower()
    if user_input == 'n':
        print("配置已取消")
        return None

    # 保存配置
    config = {
        "项目根目录": str(project_root),
        "码表路径": str(mabiao_path),
        "码表Sheet": mabiao_sheet,
        "对接人": duijieren,
        "工作目录": str(work_dir),
        "默认": {
            "城市字段": ["城市", "城市名称", "注册城市", "所属城市", "服务城市", "city_name"],
            "品牌字段": ["品牌", "品牌名称", "商家", "商家名称", "合作品牌", "合作商家"],
            "处理sheet": []
        },
        "特定配置": []
    }

    # 确保目录存在
    CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)

    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        yaml.dump(config, f, allow_unicode=True, default_flow_style=False)

    print(f"\n✅ 配置已保存到: {CONFIG_PATH}")
    return config


def load_config():
    """加载配置文件"""
    if CONFIG_PATH.exists():
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return yaml.safe_load(f)
    return None


def get_file_config(config, filename):
    """根据文件名获取配置（优先匹配特定配置）"""
    if not config:
        return {
            "城市字段": ["城市", "城市名称", "注册城市", "所属城市", "服务城市", "city_name"],
            "品牌字段": ["品牌", "品牌名称", "商家", "商家名称", "合作品牌", "合作商家"],
            "对接人": config.get("对接人", "LWL") if config else "LWL",
            "保留sheet": []
        }

    # 检查特定配置
    specific_configs = config.get("特定配置", [])
    for cfg in specific_configs:
        if cfg.get("匹配") and cfg["匹配"] in filename:
            result = config.get("默认", {}).copy()
            result.update(cfg)
            # 确保保留sheet字段存在
            if "保留sheet" not in result:
                result["保留sheet"] = []
            return result

    # 返回默认配置
    default = config.get("默认", {
        "城市字段": ["城市", "城市名称", "注册城市", "所属城市", "服务城市", "city_name"],
        "品牌字段": ["品牌", "品牌名称", "商家", "商家名称", "合作品牌", "合作商家"],
        "对接人": config.get("对接人", "LWL")
    })
    if "保留sheet" not in default:
        default["保留sheet"] = []
    return default


def load_mabiao(mabiao_path, target_duijieren, sheet_name=None):
    """加载码表，返回品牌+城市到运营主体的映射

    Args:
        mabiao_path: 码表文件路径
        target_duijieren: 对接人，可以是：
            - 单个名字（如 "LWL"）
            - 多个名字（逗号分隔，如 "LWL,张三"）
            - "全部" 表示所有对接人
        sheet_name: 码表所在sheet名称或索引，None表示第一个sheet
    """
    print(f"加载码表: {mabiao_path}")
    if sheet_name:
        print(f"  Sheet: {sheet_name}")
    df = pd.read_excel(mabiao_path, sheet_name=sheet_name if sheet_name else 0)

    # 检查必需列
    required_cols = ["品牌", "城市"]
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        # 尝试查找替代列名
        col_mapping = {}
        for col in df.columns:
            if "品牌" in str(col):
                col_mapping["品牌"] = col
            if "城市" in str(col):
                col_mapping["城市"] = col
            if "运营主体" in str(col):
                col_mapping["运营主体"] = col
            if "对接人" in str(col):
                col_mapping["对接人"] = col

        if "品牌" not in col_mapping or "城市" not in col_mapping:
            print(f"❌ 码表缺少必需列: 品牌、城市")
            print(f"   当前列: {df.columns.tolist()}")
            return {}, []

    # 识别实际列名
    brand_col = None
    city_col = None
    zhuti_col = None
    duijieren_col = None

    for col in df.columns:
        col_str = str(col)
        if "品牌" in col_str and brand_col is None:
            brand_col = col
        if "城市" in col_str and city_col is None:
            city_col = col
        if "运营主体" in col_str and zhuti_col is None:
            zhuti_col = col
        if "对接人" in col_str and duijieren_col is None:
            duijieren_col = col

    if not all([brand_col, city_col, zhuti_col, duijieren_col]):
        print(f"❌ 码表缺少必需列")
        print(f"   品牌: {brand_col}, 城市: {city_col}, 运营主体: {zhuti_col}, 对接人: {duijieren_col}")
        return {}, []

    print(f"  识别列: 品牌={brand_col}, 城市={city_col}, 运营主体={zhuti_col}, 对接人={duijieren_col}")

    # 解析对接人
    if target_duijieren == "全部":
        df_filtered = df
        duijieren_list = df[duijieren_col].dropna().unique().tolist()
        print(f"拆出所有对接人: {duijieren_list}")
    else:
        # 支持逗号分隔的多个对接人
        duijieren_list = [x.strip() for x in target_duijieren.split(",")]
        df_filtered = df[df[duijieren_col].isin(duijieren_list)]

    if len(df_filtered) == 0:
        print(f"⚠️ 码表中没有对接人为 [{target_duijieren}] 的记录")
        return {}, []

    # 创建 (品牌, 城市) -> 运营主体 的映射
    mapping = {}
    for _, row in df_filtered.iterrows():
        key = (row[brand_col], row[city_col])
        mapping[key] = row[zhuti_col]

    # 获取运营主体列表
    zhuti_list = df_filtered[zhuti_col].unique().tolist()

    print(f"对接人 {duijieren_list} 的运营主体: {zhuti_list}")
    print(f"映射记录数: {len(mapping)}")

    return mapping, zhuti_list


def detect_header_rows(ws, df, city_fields, brand_fields):
    """检测表头行数

    检测逻辑：
    1. 遍历前几行，查找包含城市或品牌字段名的行
    2. 包含关键字段名的行即为表头的最后一行
    3. 同时考虑合并单元格的跨行情况
    """
    # 先检查合并单元格跨行情况
    merge_max_row = 1
    for merged_range in ws.merged_cells.ranges:
        if merged_range.max_row > merged_range.min_row:
            merge_max_row = max(merge_max_row, merged_range.max_row)

    # 检查前5行，查找包含关键字段的行
    max_check_rows = min(5, len(df))
    last_header_row = 1

    for row_idx in range(max_check_rows):
        row_values = df.iloc[row_idx].tolist()
        for col_name in row_values:
            if col_name and str(col_name).strip() in city_fields + brand_fields:
                last_header_row = row_idx + 1  # 表头到此行结束
                break

    # 取合并单元格跨行和关键字段检测的最大值
    return max(last_header_row, merge_max_row)


def get_header_merges(ws, header_rows):
    """获取表头区域的合并单元格信息"""
    header_merges = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= header_rows:
            header_merges.append({
                'min_row': merged_range.min_row,
                'max_row': merged_range.max_row,
                'min_col': merged_range.min_col,
                'max_col': merged_range.max_col
            })
    return header_merges


def apply_cell_style(ws, header_rows):
    """应用单元格样式：字体、对齐、列宽"""
    # 统一使用微软雅黑字体（Windows/macOS 通用）
    font = Font(name='微软雅黑', size=10)
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows():
        for cell in row:
            # 跳过合并单元格
            if isinstance(cell, MergedCell):
                continue
            cell.font = font
            cell.alignment = alignment
            cell.border = thin_border

    # 使用列索引计算列宽，避免 MergedCell 问题
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value and not isinstance(cell, MergedCell):
                    # 跨平台编码处理
                    try:
                        cell_length = len(str(cell.value).encode('gbk'))
                    except UnicodeEncodeError:
                        cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = max(adjusted_width, 8)


def process_excel(file_path, mapping, zhuti_list, file_config):
    """处理单个Excel文件，保留原表头格式和数据格式"""
    print(f"\n处理文件: {file_path.name}")
    print(f"使用配置: 城市字段={file_config.get('城市字段')}, 品牌字段={file_config.get('品牌字段')}")

    # data_only=True 获取计算后的值
    wb_values = load_workbook(file_path, data_only=True)
    # data_only=False 保留格式信息
    wb_formats = load_workbook(file_path, data_only=False)
    sheet_names = wb_values.sheetnames

    process_sheets = file_config.get("处理sheet", [])
    keep_sheets = file_config.get("保留sheet", [])

    if process_sheets:
        sheet_names = [s for s in sheet_names if s in process_sheets]
        print(f"指定处理sheet: {process_sheets}，实际处理: {sheet_names}")
    else:
        print(f"Sheet列表: {sheet_names}")

    if keep_sheets:
        print(f"保留sheet（不拆分）: {keep_sheets}")

    zhuti_data = {zhuti: {} for zhuti in zhuti_list}

    # 收集保留sheet的数据（原样复制，不拆分）
    kept_sheets_data = {}

    for sheet_name in sheet_names:
        ws_values = wb_values[sheet_name]
        ws_formats = wb_formats[sheet_name]
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        print(f"\n  Sheet: {sheet_name}, 行数: {len(df)}")

        if len(df) == 0:
            continue

        # 如果是保留sheet，原样保存
        if sheet_name in keep_sheets:
            print(f"    保留sheet，原样复制")
            kept_sheets_data[sheet_name] = {
                'ws_values': ws_values,
                'ws_formats': ws_formats,
                'df': df
            }
            continue

        # 先获取字段配置，用于检测表头
        city_fields = file_config.get("城市字段", ["城市", "城市名称", "注册城市", "所属城市", "city_name"])
        brand_fields = file_config.get("品牌字段", ["品牌", "品牌名称", "商家", "商家名称", "合作品牌", "合作商家"])

        header_rows = detect_header_rows(ws_values, df, city_fields, brand_fields)
        print(f"    检测到表头行数: {header_rows}")

        header_merges = get_header_merges(ws_values, header_rows)
        headers = df.iloc[:header_rows].values.tolist()

        # 收集表头单元格格式
        header_formats = []
        for row_idx in range(1, header_rows + 1):
            row_formats = []
            for col_idx in range(1, ws_formats.max_column + 1):
                cell = ws_formats.cell(row=row_idx, column=col_idx)
                row_formats.append(cell.number_format)
            header_formats.append(row_formats)

        data_df = df.iloc[header_rows:].reset_index(drop=True)
        if len(data_df) == 0:
            continue

        # 检查所有表头行，查找品牌和城市列
        brand_col_idx = None
        city_col_idx = None

        # 遍历所有表头行查找品牌和城市列
        for header_row_idx in range(header_rows):
            header_row = df.iloc[header_row_idx].tolist()
            for idx, col_name in enumerate(header_row):
                col_str = str(col_name).strip() if col_name else ""
                if col_str in brand_fields and brand_col_idx is None:
                    brand_col_idx = idx
                    print(f"    在表头第{header_row_idx + 1}行找到品牌列: 索引{idx}")
                elif col_str in city_fields and city_col_idx is None:
                    city_col_idx = idx
                    print(f"    在表头第{header_row_idx + 1}行找到城市列: 索引{idx}")

        if brand_col_idx is None or city_col_idx is None:
            print(f"    ⚠️ 缺少品牌或城市列，跳过此sheet")
            print(f"    表头内容: {df.iloc[:header_rows].values.tolist()}")
            continue

        print(f"    最终品牌列索引: {brand_col_idx}, 城市列索引: {city_col_idx}")

        # 收集数据行和格式
        for excel_row_idx, (_, row) in enumerate(data_df.iterrows()):
            brand = row.iloc[brand_col_idx] if brand_col_idx < len(row) else None
            city = row.iloc[city_col_idx] if city_col_idx < len(row) else None

            if pd.isna(brand) or pd.isna(city):
                continue

            key = (str(brand), str(city))
            if key in mapping:
                zhuti = mapping[key]
                if sheet_name not in zhuti_data[zhuti]:
                    zhuti_data[zhuti][sheet_name] = {
                        'headers': headers,
                        'header_rows': header_rows,
                        'header_merges': header_merges,
                        'header_formats': header_formats,
                        'data_rows': [],
                        'data_formats': []
                    }

                # 保存数据值
                zhuti_data[zhuti][sheet_name]['data_rows'].append(row.values.tolist())

                # 保存数据格式（从原文件的对应行获取）
                actual_row = header_rows + excel_row_idx + 1  # Excel行号（1-based）
                row_formats = []
                for col_idx in range(1, ws_formats.max_column + 1):
                    cell = ws_formats.cell(row=actual_row, column=col_idx)
                    row_formats.append(cell.number_format)
                zhuti_data[zhuti][sheet_name]['data_formats'].append(row_formats)

    wb_values.close()
    wb_formats.close()

    for zhuti, sheets in zhuti_data.items():
        total_rows = sum(len(s.get('data_rows', [])) for s in sheets.values())
        if total_rows > 0:
            print(f"  {zhuti}: {total_rows} 行")

    return zhuti_data, kept_sheets_data


def copy_sheet_keep_original(ws_source_values, ws_source_formats, ws_target):
    """复制sheet，保留原格式"""
    # 复制值和格式
    for row_idx in range(1, ws_source_values.max_row + 1):
        for col_idx in range(1, ws_source_values.max_column + 1):
            cell_source = ws_source_values.cell(row=row_idx, column=col_idx)
            cell_format = ws_source_formats.cell(row=row_idx, column=col_idx)
            cell_target = ws_target.cell(row=row_idx, column=col_idx)

            cell_target.value = cell_source.value
            cell_target.number_format = cell_format.number_format

            # 复制样式
            if cell_format.has_style:
                cell_target.font = copy(cell_format.font)
                cell_target.border = copy(cell_format.border)
                cell_target.fill = copy(cell_format.fill)
                cell_target.number_format = cell_format.number_format
                cell_target.protection = copy(cell_format.protection)
                cell_target.alignment = copy(cell_format.alignment)

    # 复制合并单元格
    for merged_range in ws_source_values.merged_cells.ranges:
        ws_target.merge_cells(str(merged_range))

    # 复制列宽
    for col_letter, col_dim in ws_source_values.column_dimensions.items():
        ws_target.column_dimensions[col_letter].width = col_dim.width

    # 复制行高
    for row_num, row_dim in ws_source_values.row_dimensions.items():
        ws_target.row_dimensions[row_num].height = row_dim.height


def save_split_files(zhuti_data, kept_sheets_data, original_name, output_dir, file_path):
    """保存拆分后的文件，保留表头格式和数据格式"""
    saved_files = []

    for zhuti, sheets in zhuti_data.items():
        if not sheets and not kept_sheets_data:
            continue

        output_file = output_dir / f"{zhuti}_{original_name}"
        wb = Workbook()
        wb.remove(wb.active)

        # 保存拆分的sheet
        for sheet_name, sheet_data in sheets.items():
            ws = wb.create_sheet(title=sheet_name[:31])

            headers = sheet_data['headers']
            header_rows = sheet_data['header_rows']
            header_merges = sheet_data['header_merges']
            header_formats = sheet_data.get('header_formats', [])
            data_rows = sheet_data['data_rows']
            data_formats = sheet_data.get('data_formats', [])

            # 写入表头，应用格式
            for row_idx, header_row in enumerate(headers, start=1):
                for col_idx, value in enumerate(header_row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # 应用原表头格式
                    if row_idx - 1 < len(header_formats) and col_idx - 1 < len(header_formats[row_idx - 1]):
                        cell.number_format = header_formats[row_idx - 1][col_idx - 1]

            # 写入数据行，应用格式
            for row_idx, (data_row, row_formats) in enumerate(zip(data_rows, data_formats), start=header_rows + 1):
                for col_idx, value in enumerate(data_row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # 应用原数据格式
                    if col_idx - 1 < len(row_formats):
                        cell.number_format = row_formats[col_idx - 1]

            for merge_info in header_merges:
                ws.merge_cells(
                    start_row=merge_info['min_row'],
                    end_row=merge_info['max_row'],
                    start_column=merge_info['min_col'],
                    end_column=merge_info['max_col']
                )

            apply_cell_style(ws, header_rows)

        # 复制保留的sheet（原样复制到每个输出文件）
        if kept_sheets_data:
            # 重新加载原文件以获取完整格式
            wb_orig_values = load_workbook(file_path, data_only=True)
            wb_orig_formats = load_workbook(file_path, data_only=False)

            for sheet_name, _ in kept_sheets_data.items():
                ws_source_values = wb_orig_values[sheet_name]
                ws_source_formats = wb_orig_formats[sheet_name]
                ws_target = wb.create_sheet(title=sheet_name[:31])
                copy_sheet_keep_original(ws_source_values, ws_source_formats, ws_target)
                print(f"    复制保留sheet: {sheet_name}")

            wb_orig_values.close()
            wb_orig_formats.close()

        wb.save(output_file)
        saved_files.append(output_file)
        print(f"  保存: {output_file.name}")

    return saved_files


def main():
    print("=" * 60)
    print("按运营主体拆表工具")
    print("=" * 60)

    # 检查是否需要重新配置
    force_config = "--config" in sys.argv

    # 加载配置
    config = load_config()

    if not config or force_config:
        config = interactive_setup()
        if not config:
            return

    # 从配置获取路径
    project_root = Path(config["项目根目录"])
    mabiao_path = Path(config["码表路径"])
    mabiao_sheet = config.get("码表Sheet")  # 可选
    work_dir = Path(config["工作目录"])
    duijieren = config["对接人"]

    # 工作目录下的子目录
    todo_dir = work_dir / "待拆表"
    done_dir = work_dir / "已拆表"
    archive_dir = work_dir / "原表存档"

    # 检查目录
    for dir_path, dir_name in [(todo_dir, "待拆表"), (done_dir, "已拆表"), (archive_dir, "原表存档")]:
        if not dir_path.exists():
            print(f"创建目录: {dir_name}")
            dir_path.mkdir(parents=True, exist_ok=True)

    # 加载码表
    try:
        mapping, zhuti_list = load_mabiao(mabiao_path, duijieren, mabiao_sheet)
        if not mapping:
            print("\n❌ 无法加载码表映射，请检查码表文件和对接人配置")
            return
    except Exception as e:
        print(f"❌ 加载码表失败: {e}")
        return

    # 查找待拆文件
    todo_files = list(todo_dir.glob("*.xlsx")) + list(todo_dir.glob("*.xlsm"))
    if not todo_files:
        print(f"\n⚠️ 待拆目录中没有Excel文件: {todo_dir}")
        return

    print(f"\n待拆文件数: {len(todo_files)}")

    # 创建临时输出目录
    temp_dir = done_dir / f"temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    temp_dir.mkdir(exist_ok=True)

    all_saved_files = []
    processed_files = []

    for file_path in todo_files:
        file_config = get_file_config(config, file_path.name)
        zhuti_data, kept_sheets_data = process_excel(file_path, mapping, zhuti_list, file_config)
        saved = save_split_files(zhuti_data, kept_sheets_data, file_path.name, temp_dir, file_path)
        all_saved_files.extend(saved)
        if saved:
            processed_files.append(file_path)

    if not all_saved_files:
        print("\n⚠️ 没有拆分出任何文件")
        temp_dir.rmdir()
        return

    # 打包 - 使用原文件名命名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    original_name = processed_files[0].stem  # 去掉.xlsx后缀
    zip_name = f"{original_name}__{timestamp}.zip"
    zip_path = done_dir / zip_name

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_path in all_saved_files:
            zf.write(file_path, file_path.name)

    print(f"\n✅ 打包完成: {zip_path}")

    # 清理临时文件
    for file_path in all_saved_files:
        file_path.unlink()
    temp_dir.rmdir()

    # 原表存档
    for file_path in processed_files:
        archive_path = archive_dir / f"{timestamp}_{file_path.name}"
        file_path.rename(archive_path)
        print(f"原表存档: {archive_path.name}")

    print(f"\n拆分文件数: {len(all_saved_files)}")
    print("=" * 60)


if __name__ == "__main__":
    main()